#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Lead Scoring System - score_leads.py
============================================================
Tải dữ liệu từ Google Sheets hoặc file cục bộ, chấm điểm và phân loại
khách hàng tiềm năng bất động sản theo tiêu chí cho sẵn, xuất Excel chuyên nghiệp.
"""

import os
import sys
import re
import csv
import io
import argparse
from datetime import datetime
from pathlib import Path
import unicodedata

# Cấu hình UTF-8 cho console Windows
if sys.platform.startswith("win"):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except AttributeError:
        pass

# ---------------------------------------------------------------------------
# Tự động cài đặt thư viện thiếu (Self-healing imports)
# ---------------------------------------------------------------------------
def install_and_import(package, import_name=None):
    if import_name is None:
        import_name = package
    try:
        return __import__(import_name)
    except ImportError:
        import subprocess
        print(f"⚠️ Thiếu thư viện '{package}', đang cài đặt tự động...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", package])
            print(f"✅ Đã cài đặt thành công '{package}'")
            return __import__(import_name)
        except Exception as e:
            print(f"❌ Không thể cài đặt '{package}': {e}")
            return None

requests = install_and_import("requests")
openpyxl = install_and_import("openpyxl")
# Optional gspread for advanced writeback
gspread = install_and_import("gspread")

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Cấu hình & Từ khóa mặc định
# ---------------------------------------------------------------------------
BASE_SCORE = 50
VIP_BONUS = 50
JUNK_PENALTY = -50

VIP_KEYWORDS = {
    "large_budget_terms": ["tài chính mạnh", "tai chinh manh", "không thành vấn đề", "khong thanh van de", "tài chính lớn", "tai chinh lon"],
    "luxury_types": ["biệt thự đơn lập", "biet thu don lap", "penthouse", "shophouse mặt đường lớn", "shophouse mat duong lon", "shophouse mặt đường", "quỹ đất công nghiệp", "quy dat cong nghiep", "sàn văn phòng diện tích lớn", "san van phong dien tich lon", "shophouse"],
    "prime_locations": ["quận 1", "quan 1", "ven sông", "ven song", "vinhomes ocean park", "vinhomes oceanpark", "phú mỹ hưng", "phu my hung"],
    "vip_profiles": ["chủ doanh nghiệp", "chu doanh nghiep", "nhà đầu tư chuyên nghiệp", "nha dau tu chuyen nghiep", "mua sỉ", "mua si", "mua số lượng lớn", "mua so luong lon"],
    "urgency_transparency": ["pháp lý chuẩn 100%", "phap ly chuan 100%", "sổ hồng riêng", "so hong rieng", "gặp trực tiếp chủ đầu tư", "gap truc tiep chu dau tu"]
}

JUNK_KEYWORDS = {
    "no_need": ["nhầm số", "nham so", "không có nhu cầu", "khong co nhu cau", "dữ liệu cũ", "du lieu cu", "nhầm ngành", "nham nganh"],
    "uncooperative": ["hỏi giá cho vui", "hoi gia cho vui", "chưa có ý định mua", "chua co y dinh mua", "thái độ không hợp tác", "thai do khong hop tac"],
    "spam_adv": ["bảo hiểm", "bao hiem", "vay vốn", "vay von", "mời chào dịch vụ", "moi chao dich vu", "quảng cáo", "quang cao"],
    "contact_issue": ["thuê bao", "thue bao", "không bắt máy", "khong bat may", "không nhấc máy", "khong nhac may", "gọi nhiều lần", "goi nhieu lan", "không phản hồi zalo", "khong phan hoi zalo"]
}

# ---------------------------------------------------------------------------
# Tiền xử lý văn bản tiếng Việt
# ---------------------------------------------------------------------------
def remove_accents(input_str):
    """Bỏ dấu tiếng Việt."""
    if not input_str:
        return ""
    nfkd_form = unicodedata.normalize('NFKD', input_str)
    res = "".join([c for c in nfkd_form if not unicodedata.combining(c)])
    # Thay thế chữ đ/Đ thành d/D
    res = res.replace('đ', 'd').replace('Đ', 'D')
    return res

def normalize_text(text):
    """Trả về văn bản viết thường có dấu và không dấu để so khớp."""
    if not text:
        return "", ""
    text_lower = text.strip().lower()
    text_no_accents = remove_accents(text_lower)
    return text_lower, text_no_accents

# ---------------------------------------------------------------------------
# Logic Chấm Điểm
# ---------------------------------------------------------------------------
def evaluate_lead(row):
    """
    Đánh giá chất lượng lead dựa trên mô tả nhu cầu.
    Trả về: (final_score, category, reasoning_details, vip_matched, junk_matched)
    """
    nhu_cau = row.get("nhu_cau_mo_ta", "")
    if not nhu_cau:
        return BASE_SCORE, "Bình thường", "Không có mô tả nhu cầu để đánh giá.", [], []
        
    text_lower, text_no_acc = normalize_text(nhu_cau)
    
    vip_matched = []
    junk_matched = []
    
    # --- 1. KIỂM TRA TIÊU CHÍ CỘNG 50 ĐIỂM (VIP) ---
    
    # a. Ngân sách lớn (>= 20 tỷ hoặc từ khóa VIP ngân sách)
    # Khớp số tỷ: ví dụ "20 tỷ", "20.5 tỷ", "100 ty"
    budget_found = False
    budget_reason = ""
    matches = re.findall(r"(\d+(?:[.,]\d+)?)\s*(?:tỷ|ty)", text_lower)
    for num_str in matches:
        try:
            val = float(num_str.replace(",", "."))
            if val >= 20.0:
                budget_found = True
                budget_reason = f"Ngân sách lớn ({num_str} tỷ)"
                break
        except ValueError:
            continue
            
    for kw in VIP_KEYWORDS["large_budget_terms"]:
        if kw in text_lower or kw in text_no_acc:
            budget_found = True
            budget_reason = f"Ngân sách lớn (Từ khóa: '{kw}')"
            break
            
    if budget_found:
        vip_matched.append(budget_reason)
        
    # b. Loại hình cao cấp
    for kw in VIP_KEYWORDS["luxury_types"]:
        if kw in text_lower or kw in text_no_acc:
            vip_matched.append(f"Loại hình cao cấp ({kw.capitalize()})")
            break
            
    # c. Vị trí đắc địa
    for kw in VIP_KEYWORDS["prime_locations"]:
        # Chú ý so khớp Quận 1 chính xác hơn để tránh dính "quận 10", "quận 11"
        if kw in ["quận 1", "quan 1"]:
            if re.search(r"\b(quận 1|quan 1|q1|q\.1)\b", text_lower):
                vip_matched.append("Vị trí đắc địa (Quận 1)")
                break
        elif kw in text_lower or kw in text_no_acc:
            vip_matched.append(f"Vị trí đắc địa ({kw.capitalize()})")
            break
            
    # d. Đối tượng VIP
    for kw in VIP_KEYWORDS["vip_profiles"]:
        if kw in text_lower or kw in text_no_acc:
            vip_matched.append(f"Đối tượng khách hàng ({kw.capitalize()})")
            break
            
    # e. Tính cấp thiết & Minh bạch
    for kw in VIP_KEYWORDS["urgency_transparency"]:
        if kw in text_lower or kw in text_no_acc:
            vip_matched.append(f"Tính minh bạch/Cấp thiết ({kw.capitalize()})")
            break
            
    # --- 2. KIỂM TRA TIÊU CHÍ TRỪ 50 ĐIỂM (JUNK) ---
    
    # a. Yêu cầu phi thực tế (Giá thấp vô lý ở khu vực trung tâm/Quận 1)
    unrealistic_found = False
    unrealistic_reason = ""
    # Kiểm tra Quận 1 giá rẻ
    has_q1 = bool(re.search(r"\b(quận 1|quan 1|q1|q\.1)\b", text_lower))
    if has_q1:
        # Tìm xem có đề cập số tỷ thấp (<= 2 tỷ) hoặc vài trăm triệu
        low_price_kws = ["1 tỷ", "1 ty", "2 tỷ", "2 ty", "1-2 tỷ", "1-2 ty", "dưới 2 tỷ", "dưới 2 ty", "vài trăm triệu", "vai tram trieu", "mấy trăm triệu", "may tram trieu"]
        for lp in low_price_kws:
            if lp in text_lower or lp in text_no_acc:
                unrealistic_found = True
                unrealistic_reason = f"Yêu cầu phi thực tế (Quận 1 giá rẻ: '{lp}')"
                break
        if not unrealistic_found:
            matches_q1_prices = re.findall(r"(\d+(?:[.,]\d+)?)\s*(?:tỷ|ty)", text_lower)
            for num_str in matches_q1_prices:
                try:
                    val = float(num_str.replace(",", "."))
                    if val <= 2.5:
                        unrealistic_found = True
                        unrealistic_reason = f"Yêu cầu phi thực tế (Quận 1 giá quá thấp: {num_str} tỷ)"
                        break
                except ValueError:
                    continue
                    
    # Kiểm tra trung tâm có sân vườn/hồ bơi giá vài trăm triệu
    has_central = any(kw in text_lower or kw in text_no_acc for kw in ["trung tâm", "trung tam", "nội thành", "noi thanh"])
    has_luxury_feature = any(kw in text_lower or kw in text_no_acc for kw in ["sân vườn", "san vuon", "hồ bơi", "ho boi"])
    if has_central and has_luxury_feature:
        low_price_kws = ["vài trăm triệu", "vai tram trieu", "mấy trăm triệu", "may tram trieu", "vài trăm", "vai tram"]
        for lp in low_price_kws:
            if lp in text_lower or lp in text_no_acc:
                unrealistic_found = True
                unrealistic_reason = "Yêu cầu phi thực tế (Nhà trung tâm có sân vườn/hồ bơi giá vài trăm triệu)"
                break
                
    if unrealistic_found:
        junk_matched.append(unrealistic_reason)
        
    # b. Không có nhu cầu
    for kw in JUNK_KEYWORDS["no_need"]:
        if kw in text_lower or kw in text_no_acc:
            junk_matched.append(f"Không có nhu cầu ({kw.capitalize()})")
            break
            
    # c. Khách hàng không thiện chí
    for kw in JUNK_KEYWORDS["uncooperative"]:
        if kw in text_lower or kw in text_no_acc:
            junk_matched.append(f"Không thiện chí ({kw.capitalize()})")
            break
            
    # d. Spam/Quảng cáo
    for kw in JUNK_KEYWORDS["spam_adv"]:
        if kw in text_lower or kw in text_no_acc:
            junk_matched.append(f"Spam/Quảng cáo ({kw.capitalize()})")
            break
            
    # e. Thông tin liên lạc lỗi
    for kw in JUNK_KEYWORDS["contact_issue"]:
        if kw in text_lower or kw in text_no_acc:
            junk_matched.append(f"Liên lạc lỗi ({kw.capitalize()})")
            break
            
    # --- 3. TÍNH ĐIỂM CUỐI CÙNG ---
    vip_points = VIP_BONUS if vip_matched else 0
    junk_points = JUNK_PENALTY if junk_matched else 0
    
    final_score = BASE_SCORE + vip_points + junk_points
    # Giới hạn trong khoảng [0, 100]
    final_score = max(0, min(100, final_score))
    
    # Phân loại và giải trình lý do
    if final_score >= 100:
        category = "VIP"
        reason = f"Cộng 50 điểm (Khách VIP): Đạt tiêu chí. Chi tiết: {', '.join(vip_matched)}"
        if junk_matched:
            reason += f" | Tuy nhiên có dấu hiệu tiêu cực: {', '.join(junk_matched)}"
    elif final_score <= 0:
        category = "Rác"
        reason = f"Trừ 50 điểm (Khách Rác): Phát hiện dấu hiệu tiêu cực. Chi tiết: {', '.join(junk_matched)}"
        if vip_matched:
            reason += f" | Có chứa từ khóa VIP: {', '.join(vip_matched)}"
    else:
        category = "Bình thường"
        if vip_matched and junk_matched:
            reason = f"Giữ nguyên 50 điểm (Bình thường): Giao thoa cả tiêu chí VIP và tiêu chí Rác. VIP: {', '.join(vip_matched)} | Rác: {', '.join(junk_matched)}"
        else:
            reason = "Giữ nguyên 50 điểm (Bình thường): Khách hàng tầm trung hoặc nhu cầu thực tế cần tư vấn."
            
    return final_score, category, reason, vip_matched, junk_matched

# ---------------------------------------------------------------------------
# Đọc Dữ Liệu
# ---------------------------------------------------------------------------
def load_data(source_url):
    """
    Tải dữ liệu từ URL Google Sheets (dưới dạng export CSV) hoặc file cục bộ.
    """
    if source_url.startswith("http://") or source_url.startswith("https://"):
        # Chuyển link edit Google Sheet thành link export CSV
        if "docs.google.com/spreadsheets" in source_url:
            # Lấy Spreadsheet ID
            match = re.search(r"/d/([a-zA-Z0-9-_]+)", source_url)
            if match:
                sheet_id = match.group(1)
                source_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
                print(f"ℹ️ Đã chuyển URL Google Sheet thành URL xuất CSV: {source_url}")
                
        print(f"📥 Đang tải dữ liệu từ Google Sheet...")
        headers = {'User-Agent': 'Mozilla/5.0'}
        response = requests.get(source_url, headers=headers)
        if response.status_code != 200:
            raise Exception(f"Không thể tải file từ Google Sheet (HTTP Status Code: {response.status_code})")
            
        csv_data = response.content.decode('utf-8')
        reader = csv.DictReader(io.StringIO(csv_data))
        rows = list(reader)
        return rows
    else:
        # File cục bộ
        path = Path(source_url)
        if not path.exists():
            raise FileNotFoundError(f"Không tìm thấy file nguồn tại: {source_url}")
            
        print(f"📖 Đang đọc file cục bộ: {source_url}")
        if path.suffix.lower() == '.csv':
            with open(path, mode='r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                return list(reader)
        elif path.suffix.lower() in ['.xlsx', '.xls']:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows = []
            for r_idx in range(2, ws.max_row + 1):
                row_dict = {}
                for c_idx, h in enumerate(headers, 1):
                    if h:
                        row_dict[h] = ws.cell(row=r_idx, column=c_idx).value
                rows.append(row_dict)
            return rows
        else:
            raise ValueError("Định dạng file không hỗ trợ. Chỉ hỗ trợ .csv, .xlsx, .xls hoặc link Google Sheet.")

# ---------------------------------------------------------------------------
# Tạo File Báo Cáo Excel Chuyên Nghiệp (Theme Deep Navy)
# ---------------------------------------------------------------------------
def create_excel_report(scored_leads, integrity_log, output_path):
    wb = openpyxl.Workbook()
    # Xoá sheet mặc định
    wb.remove(wb.active)
    
    # -----------------------------------------------------------------------
    # Sheet 1: Summary_Dashboard
    # -----------------------------------------------------------------------
    ws_dash = wb.create_sheet(title="Summary_Dashboard")
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Định dạng Font & Màu
    FONT_FAMILY = "Segoe UI"
    font_title = Font(name=FONT_FAMILY, size=16, bold=True, color="1F4E79")
    font_subtitle = Font(name=FONT_FAMILY, size=10, italic=True, color="595959")
    font_section = Font(name=FONT_FAMILY, size=12, bold=True, color="1F4E79")
    font_kpi_val = Font(name=FONT_FAMILY, size=22, bold=True, color="1F4E79")
    font_kpi_lbl = Font(name=FONT_FAMILY, size=9, bold=True, color="595959")
    
    fill_navy_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_kpi = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    # Viền mỏng
    thin_border = Side(border_style="thin", color="D9D9D9")
    double_border = Side(border_style="double", color="1F4E79")
    thick_border = Side(border_style="medium", color="1F4E79")
    
    border_kpi = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thick_border)
    border_cell = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    border_total = Border(top=thin_border, bottom=double_border)
    
    # Tiêu đề Dashboard
    ws_dash["A1"] = "BÁO CÁO PHÂN TÍCH CHẤT LƯỢNG KHÁCH HÀNG TIỀM NĂNG"
    ws_dash["A1"].font = font_title
    ws_dash["A2"] = f"Hệ thống chấm điểm tự động • Thời gian xuất báo cáo: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws_dash["A2"].font = font_subtitle
    ws_dash.row_dimensions[1].height = 25
    ws_dash.row_dimensions[2].height = 18
    
    # Tính toán các chỉ số KPI
    total_leads = len(scored_leads)
    vip_leads = sum(1 for r in scored_leads if r["__score"][1] == "VIP")
    normal_leads = sum(1 for r in scored_leads if r["__score"][1] == "Bình thường")
    junk_leads = sum(1 for r in scored_leads if r["__score"][1] == "Rác")
    avg_score = sum(r["__score"][0] for r in scored_leads) / total_leads if total_leads > 0 else 0
    
    # Thiết kế KPI Cards (B4:F5)
    kpi_cols = [("B", "TỔNG SỐ KHÁCH", total_leads, "#,##0"), 
                ("C", "KHÁCH VIP (+50)", vip_leads, "#,##0"), 
                ("D", "BÌNH THƯỜNG (50đ)", normal_leads, "#,##0"), 
                ("E", "KHÁCH RÁC (-50)", junk_leads, "#,##0"), 
                ("F", "ĐIỂM TRUNG BÌNH", avg_score, "0.0")]
                
    for col, lbl, val, fmt in kpi_cols:
        lbl_cell = ws_dash[f"{col}4"]
        val_cell = ws_dash[f"{col}5"]
        
        lbl_cell.value = lbl
        lbl_cell.font = font_kpi_lbl
        lbl_cell.fill = fill_kpi
        lbl_cell.alignment = Alignment(horizontal="center", vertical="center")
        lbl_cell.border = Border(left=thin_border, right=thin_border, top=thin_border)
        
        val_cell.value = val
        val_cell.font = font_kpi_val
        val_cell.fill = fill_kpi
        val_cell.number_format = fmt
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        val_cell.border = Border(left=thin_border, right=thin_border, bottom=thick_border)
        
    ws_dash.row_dimensions[4].height = 18
    ws_dash.row_dimensions[5].height = 32
    
    # Bảng phân tích chi tiết (B7:C12)
    ws_dash["B7"] = "BẢNG PHÂN PHỐI CƠ CẤU KHÁCH HÀNG"
    ws_dash["B7"].font = font_section
    ws_dash.merge_cells("B7:C7")
    ws_dash.row_dimensions[7].height = 20
    
    ws_dash["B8"] = "Nhóm Khách Hàng"
    ws_dash["B8"].font = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
    ws_dash["B8"].fill = fill_navy_header
    ws_dash["B8"].alignment = Alignment(horizontal="left", vertical="center")
    
    ws_dash["C8"] = "Số Lượng"
    ws_dash["C8"].font = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
    ws_dash["C8"].fill = fill_navy_header
    ws_dash["C8"].alignment = Alignment(horizontal="right", vertical="center")
    ws_dash.row_dimensions[8].height = 22
    
    categories_data = [("VIP (100đ)", vip_leads),
                       ("Bình thường (50đ)", normal_leads),
                       ("Rác (0đ)", junk_leads)]
                       
    for idx, (cat, val) in enumerate(categories_data, 9):
        c_lbl = ws_dash[f"B{idx}"]
        c_val = ws_dash[f"C{idx}"]
        
        c_lbl.value = cat
        c_lbl.font = Font(name=FONT_FAMILY, size=10)
        c_lbl.border = border_cell
        c_lbl.alignment = Alignment(horizontal="left", vertical="center")
        
        c_val.value = val
        c_val.font = Font(name=FONT_FAMILY, size=10)
        c_val.number_format = "#,##0"
        c_val.border = border_cell
        c_val.alignment = Alignment(horizontal="right", vertical="center")
        
        if idx % 2 == 0:
            c_lbl.fill = fill_zebra
            c_val.fill = fill_zebra
            
    # Hàng tổng
    total_idx = 12
    c_lbl = ws_dash[f"B{total_idx}"]
    c_val = ws_dash[f"C{total_idx}"]
    c_lbl.value = "Tổng cộng"
    c_lbl.font = Font(name=FONT_FAMILY, size=10, bold=True)
    c_lbl.border = border_total
    c_lbl.alignment = Alignment(horizontal="left", vertical="center")
    
    c_val.value = f"=SUM(C9:C11)"
    c_val.font = Font(name=FONT_FAMILY, size=10, bold=True)
    c_val.border = border_total
    c_val.alignment = Alignment(horizontal="right", vertical="center")
    
    # Thiết lập độ rộng cột B và C cho đẹp
    ws_dash.column_dimensions["A"].width = 3
    ws_dash.column_dimensions["B"].width = 25
    ws_dash.column_dimensions["C"].width = 15
    ws_dash.column_dimensions["D"].width = 18
    ws_dash.column_dimensions["E"].width = 18
    ws_dash.column_dimensions["F"].width = 18
    
    # Vẽ Biểu đồ cơ cấu khách hàng (Pie Chart)
    try:
        from openpyxl.chart import PieChart, Reference
        pie = PieChart()
        labels = Reference(ws_dash, min_col=2, min_row=9, max_row=11)
        data = Reference(ws_dash, min_col=3, min_row=8, max_row=11)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Cơ Cấu Khách Hàng Tiềm Năng"
        pie.width = 16
        pie.height = 10
        # Đặt biểu đồ tại vị trí B14
        ws_dash.add_chart(pie, "B14")
        print("📊 Đã tạo biểu đồ cơ cấu khách hàng.")
    except Exception as chart_err:
        print(f"⚠️ Không thể vẽ biểu đồ: {chart_err}")
        
    # -----------------------------------------------------------------------
    # Sheet 2: Lead_Scores_Details
    # -----------------------------------------------------------------------
    ws_detail = wb.create_sheet(title="Lead_Scores_Details")
    ws_detail.views.sheetView[0].showGridLines = True
    
    headers = [
        "Mã Lead", "Tên Khách Hàng", "Số Điện Thoại", "Mô Tả Nhu Cầu",
        "Điểm Cơ Bản", "Điểm Cộng (VIP)", "Điểm Trừ (Rác)", "Điểm Cuối Cùng",
        "Phân Loại", "Giải Trình Lý Do Chấm Điểm"
    ]
    
    # Write header row
    for col_idx, h in enumerate(headers, 1):
        cell = ws_detail.cell(row=1, column=col_idx, value=h)
        cell.font = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
        cell.fill = fill_navy_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_cell
    ws_detail.row_dimensions[1].height = 28
    
    # Màu sắc định dạng cho từng nhóm
    fill_vip = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # light green
    font_vip = Font(name=FONT_FAMILY, size=10, bold=True, color="375623")
    
    fill_junk = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # light red
    font_junk = Font(name=FONT_FAMILY, size=10, bold=True, color="C65911")
    
    fill_normal_badge = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # light blue
    font_normal_badge = Font(name=FONT_FAMILY, size=10, bold=True, color="1F4E79")
    
    for r_idx, lead in enumerate(scored_leads, 2):
        score, cat, reason, vip_m, junk_m = lead["__score"]
        
        ws_detail.cell(row=r_idx, column=1, value=lead.get("id", f"L{r_idx-1}")).alignment = Alignment(horizontal="center")
        ws_detail.cell(row=r_idx, column=2, value=lead.get("ten_khach", "Ẩn danh"))
        ws_detail.cell(row=r_idx, column=3, value=lead.get("sdt", "N/A")).alignment = Alignment(horizontal="center")
        
        desc_cell = ws_detail.cell(row=r_idx, column=4, value=lead.get("nhu_cau_mo_ta", ""))
        desc_cell.alignment = Alignment(wrap_text=True)
        
        # Điểm số
        ws_detail.cell(row=r_idx, column=5, value=BASE_SCORE).number_format = "#,##0"
        ws_detail.cell(row=r_idx, column=6, value=VIP_BONUS if vip_m else 0).number_format = "#,##0"
        ws_detail.cell(row=r_idx, column=7, value=JUNK_PENALTY if junk_m else 0).number_format = "#,##0"
        
        score_cell = ws_detail.cell(row=r_idx, column=8, value=score)
        score_cell.font = Font(name=FONT_FAMILY, size=10, bold=True)
        score_cell.number_format = "#,##0"
        score_cell.alignment = Alignment(horizontal="center")
        
        # Phân loại
        cat_cell = ws_detail.cell(row=r_idx, column=9, value=cat)
        cat_cell.alignment = Alignment(horizontal="center")
        if cat == "VIP":
            cat_cell.fill = fill_vip
            cat_cell.font = font_vip
        elif cat == "Rác":
            cat_cell.fill = fill_junk
            cat_cell.font = font_junk
        else:
            cat_cell.fill = fill_normal_badge
            cat_cell.font = font_normal_badge
            
        # Giải trình
        ws_detail.cell(row=r_idx, column=10, value=reason).alignment = Alignment(wrap_text=True)
        
        # Border & Zebra
        for col_idx in range(1, len(headers) + 1):
            c = ws_detail.cell(row=r_idx, column=col_idx)
            c.border = border_cell
            if col_idx not in [8, 9]:
                c.font = Font(name=FONT_FAMILY, size=10)
            if r_idx % 2 == 0 and col_idx != 9: # giữ màu của ô Phân loại
                c.fill = fill_zebra
                
        ws_detail.row_dimensions[r_idx].height = 36
        
    # Auto-fit column widths
    for col in ws_detail.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        # Khống chế giới hạn chiều rộng cột
        if col_letter in ['D', 'J']:
            ws_detail.column_dimensions[col_letter].width = 45
        elif col_letter in ['A', 'C', 'E', 'F', 'G', 'H', 'I']:
            ws_detail.column_dimensions[col_letter].width = 15
        else:
            ws_detail.column_dimensions[col_letter].width = min(max_len + 3, 30)

    # -----------------------------------------------------------------------
    # Sheet 3: Data_Integrity_Log
    # -----------------------------------------------------------------------
    ws_log = wb.create_sheet(title="Data_Integrity_Log")
    ws_log.views.sheetView[0].showGridLines = True
    
    log_headers = ["Dòng Sheet", "Mã ID", "Tên Khách Hàng", "Số Điện Thoại", "Trường Bị Lỗi", "Chi Tiết Lỗi", "Hướng Xử Lý"]
    
    for col_idx, h in enumerate(log_headers, 1):
        cell = ws_log.cell(row=1, column=col_idx, value=h)
        cell.font = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
        cell.fill = fill_navy_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_cell
    ws_log.row_dimensions[1].height = 25
    
    if integrity_log:
        for r_idx, err in enumerate(integrity_log, 2):
            ws_log.cell(row=r_idx, column=1, value=err.get("row_num")).alignment = Alignment(horizontal="center")
            ws_log.cell(row=r_idx, column=2, value=err.get("id")).alignment = Alignment(horizontal="center")
            ws_log.cell(row=r_idx, column=3, value=err.get("ten_khach", "N/A"))
            ws_log.cell(row=r_idx, column=4, value=err.get("sdt", "N/A")).alignment = Alignment(horizontal="center")
            ws_log.cell(row=r_idx, column=5, value=err.get("field")).alignment = Alignment(horizontal="center")
            
            err_cell = ws_log.cell(row=r_idx, column=6, value=err.get("error_msg"))
            err_cell.font = Font(name=FONT_FAMILY, size=10, color="9C0006")
            err_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            ws_log.cell(row=r_idx, column=7, value=err.get("action"))
            
            for col_idx in range(1, len(log_headers) + 1):
                c = ws_log.cell(row=r_idx, column=col_idx)
                c.border = border_cell
                if col_idx != 6:
                    c.font = Font(name=FONT_FAMILY, size=10)
                if r_idx % 2 == 0 and col_idx != 6:
                    c.fill = fill_zebra
            ws_log.row_dimensions[r_idx].height = 24
    else:
        # Ghi một dòng thông báo sạch
        ws_log.merge_cells("A2:G2")
        empty_cell = ws_log.cell(row=2, column=1, value="🎉 Tuyệt vời! Không phát hiện lỗi chất lượng dữ liệu nào.")
        empty_cell.font = Font(name=FONT_FAMILY, size=11, bold=True, color="375623")
        empty_cell.fill = fill_vip
        empty_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_log.row_dimensions[2].height = 32
        for col_idx in range(1, len(log_headers) + 1):
            ws_log.cell(row=2, column=col_idx).border = border_cell
            
    # Auto-fit column widths for Log sheet
    for col in ws_log.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws_log.column_dimensions[col_letter].width = min(max_len + 4, 45)
        
    # Lưu file
    wb.save(output_path)
    print(f"🎉 Báo cáo Excel đã được xuất thành công ra: {output_path}")

# ---------------------------------------------------------------------------
# Cập nhật ngược lại Google Sheet bằng gspread (Tùy chọn)
# ---------------------------------------------------------------------------
def write_scores_to_gspread(scored_leads, source_url):
    if gspread is None:
        print("⚠️ Chưa cài gspread. Bỏ qua ghi đè kết quả lên Google Sheet.")
        return
        
    TOKEN_PATH = Path.home() / ".config" / "ai_audit" / "token.json"
    if not TOKEN_PATH.exists():
        print("⚠️ Không tìm thấy token OAuth tại ~/.config/ai_audit/token.json. Bỏ qua cập nhật Google Sheet.")
        return
        
    try:
        from google.oauth2.credentials import Credentials as OAuthCreds
        from google.auth.transport.requests import Request
        SCOPES = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
        
        print("🔄 Đang thử xác thực Google API...")
        creds = OAuthCreds.from_authorized_user_file(str(TOKEN_PATH), SCOPES)
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
            
        gc = gspread.authorize(creds)
        sh = gc.open_by_url(source_url)
        ws = sh.get_worksheet(0)
        
        # Đọc headers hiện tại của sheet
        headers = ws.row_values(1)
        
        # Tìm hoặc tạo cột mới
        new_cols = ["Final_Score", "Category", "Reasoning_Details"]
        col_indices = {}
        
        for col_name in new_cols:
            if col_name in headers:
                col_indices[col_name] = headers.index(col_name) + 1
            else:
                # Thêm cột mới
                headers.append(col_name)
                ws.append_row([]) # Tạo thêm cột trống nếu cần (hoặc ghi đè trực tiếp)
                col_indices[col_name] = len(headers)
                ws.update_cell(1, col_indices[col_name], col_name)
                
        print(f"📝 Đang cập nhật điểm trực tiếp lên Google Sheet '{sh.title}'...")
        
        # Chuẩn bị danh sách cập nhật hàng loạt để tối ưu API call
        updates = []
        for idx, lead in enumerate(scored_leads, 2):
            score, cat, reason, _, _ = lead["__score"]
            updates.append({
                'range': gspread.utils.rowcol_to_a1(idx, col_indices["Final_Score"]),
                'values': [[score]]
            })
            updates.append({
                'range': gspread.utils.rowcol_to_a1(idx, col_indices["Category"]),
                'values': [[cat]]
            })
            updates.append({
                'range': gspread.utils.rowcol_to_a1(idx, col_indices["Reasoning_Details"]),
                'values': [[reason]]
            })
            
        ws.batch_update(updates)
        print("✅ Cập nhật kết quả lên Google Sheet hoàn tất!")
        
    except Exception as e:
        print(f"⚠️ Gặp lỗi khi cập nhật lên Google Sheet: {e}")

# ---------------------------------------------------------------------------
# Hàm Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Chấm điểm Khách hàng Tiềm năng Bất động sản")
    parser.add_argument("--url", required=True, help="Đường dẫn Google Sheet hoặc file CSV/Excel cục bộ")
    parser.add_argument("--output", default="Lead_Scoring_Report.xlsx", help="Đường dẫn lưu file Excel báo cáo đầu ra")
    parser.add_argument("--write-back", action="store_true", help="Ghi ngược kết quả điểm số lại Google Sheet (yêu cầu OAuth)")
    args = parser.parse_args()
    
    print("=" * 60)
    print("HỆ THỐNG CHẤM ĐIỂM KHÁCH HÀNG TIỀM NĂNG BẤT ĐỘNG SẢN")
    print("=" * 60)
    
    # 1. Tải dữ liệu
    try:
        raw_rows = load_data(args.url)
        print(f"📊 Đã tải thành công {len(raw_rows)} bản ghi.")
    except Exception as e:
        print(f"❌ Lỗi tải dữ liệu: {e}")
        sys.exit(1)
        
    # 2. Làm sạch, kiểm tra tính toàn vẹn và chấm điểm
    scored_leads = []
    integrity_log = []
    
    # Set để kiểm tra trùng lặp số điện thoại
    seen_phones = {}
    
    for idx, row in enumerate(raw_rows, 2): # 1-indexed, dòng tiêu đề là 1, data bắt đầu từ dòng 2
        lead_id = row.get("id", "").strip() if row.get("id") else f"L{idx-1}"
        ten = row.get("ten_khach", "").strip() if row.get("ten_khach") else ""
        sdt = str(row.get("sdt", "")).strip() if row.get("sdt") else ""
        nhu_cau = row.get("nhu_cau_mo_ta", "").strip() if row.get("nhu_cau_mo_ta") else ""
        
        # Đồng bộ dữ liệu trong dictionary để xử lý tiếp
        row["id"] = lead_id
        row["ten_khach"] = ten if ten else "Ẩn danh"
        row["sdt"] = sdt if sdt else "N/A"
        row["nhu_cau_mo_ta"] = nhu_cau
        
        # Kiểm tra tính toàn vẹn dữ liệu
        is_valid = True
        
        if not ten:
            integrity_log.append({
                "row_num": idx,
                "id": lead_id,
                "ten_khach": "N/A",
                "sdt": sdt,
                "field": "ten_khach",
                "error_msg": "Thiếu tên khách hàng",
                "action": "Gán mặc định 'Ẩn danh'"
            })
            
        if not sdt or sdt == "N/A":
            integrity_log.append({
                "row_num": idx,
                "id": lead_id,
                "ten_khach": row["ten_khach"],
                "sdt": "Trống",
                "field": "sdt",
                "error_msg": "Thiếu số điện thoại liên lạc",
                "action": "Gán mặc định 'N/A' và bỏ qua định dạng"
            })
            is_valid = False
        else:
            # Check trùng lặp số điện thoại
            if sdt in seen_phones:
                integrity_log.append({
                    "row_num": idx,
                    "id": lead_id,
                    "ten_khach": row["ten_khach"],
                    "sdt": sdt,
                    "field": "sdt",
                    "error_msg": f"Số điện thoại trùng lặp với dòng {seen_phones[sdt]}",
                    "action": "Vẫn giữ lại và chấm điểm độc lập"
                })
            else:
                seen_phones[sdt] = idx
                
        if not nhu_cau:
            integrity_log.append({
                "row_num": idx,
                "id": lead_id,
                "ten_khach": row["ten_khach"],
                "sdt": row["sdt"],
                "field": "nhu_cau_mo_ta",
                "error_msg": "Thiếu mô tả nhu cầu bất động sản",
                "action": "Không thể chấm điểm, giữ điểm mặc định 50"
            })
            
        # Chấm điểm
        score, cat, reason, vip_m, junk_m = evaluate_lead(row)
        row["__score"] = (score, cat, reason, vip_m, junk_m)
        scored_leads.append(row)
        
    print(f"🧹 Hoàn tất làm sạch dữ liệu. Phát hiện {len(integrity_log)} điểm cảnh báo dữ liệu.")
    
    # 3. Xuất file báo cáo Excel
    create_excel_report(scored_leads, integrity_log, args.output)
    
    # 4. Ghi ngược lại Google Sheet nếu người dùng yêu cầu
    if args.write_back and args.url.startswith("http"):
        write_scores_to_gspread(scored_leads, args.url)
        
    print("\n📊 Tóm tắt kết quả phân loại khách hàng:")
    vip_c = sum(1 for r in scored_leads if r["__score"][1] == "VIP")
    normal_c = sum(1 for r in scored_leads if r["__score"][1] == "Bình thường")
    junk_c = sum(1 for r in scored_leads if r["__score"][1] == "Rác")
    print(f"   ⭐ VIP:          {vip_c} khách ({vip_c/len(scored_leads)*100:.1f}%)")
    print(f"   👥 Bình thường:  {normal_c} khách ({normal_c/len(scored_leads)*100:.1f}%)")
    print(f"   🗑️ Rác:          {junk_c} khách ({junk_c/len(scored_leads)*100:.1f}%)")
    print("=" * 60)

if __name__ == "__main__":
    main()
