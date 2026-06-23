#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
score_leads.py - Core lead scoring logic for Real Estate CRM
============================================================
Standalone module: load data, score leads, export Excel reports.
No print statements at module-level to avoid Windows encoding issues.
"""

import os
import re
import csv
import io
from datetime import datetime
from pathlib import Path
import unicodedata

import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration & Default Keywords
# ---------------------------------------------------------------------------
BASE_SCORE = 50
VIP_BONUS = 50
JUNK_PENALTY = -50

VIP_KEYWORDS = {
    "large_budget_terms": [
        "tai chinh manh", "khong thanh van de", "tai chinh lon",
        "tai chinh manh", "khong thanh van de", "tai chinh lon"
    ],
    "luxury_types": [
        "biet thu don lap", "penthouse", "shophouse mat duong lon",
        "shophouse mat duong", "quy dat cong nghiep",
        "san van phong dien tich lon", "shophouse"
    ],
    "prime_locations": [
        "quan 1", "ven song", "vinhomes ocean park", "vinhomes oceanpark",
        "phu my hung"
    ],
    "vip_profiles": [
        "chu doanh nghiep", "nha dau tu chuyen nghiep",
        "mua si", "mua so luong lon"
    ],
    "urgency_transparency": [
        "phap ly chuan 100%", "so hong rieng", "gap truc tiep chu dau tu"
    ]
}

JUNK_KEYWORDS = {
    "no_need": [
        "nham so", "khong co nhu cau", "du lieu cu", "nham nganh"
    ],
    "uncooperative": [
        "hoi gia cho vui", "chua co y dinh mua", "thai do khong hop tac"
    ],
    "spam_adv": [
        "bao hiem", "vay von", "moi chao dich vu", "quang cao"
    ],
    "contact_issue": [
        "thue bao", "khong bat may", "khong nhac may",
        "goi nhieu lan", "khong phan hoi zalo"
    ]
}

# Vietnamese diacritics map for keyword matching
VIP_KEYWORDS_VI = {
    "large_budget_terms": [
        "tai chinh manh", "khong thanh van de", "tai chinh lon"
    ],
    "luxury_types": [
        "biet thu don lap", "penthouse", "shophouse mat duong lon",
        "shophouse mat duong", "quy dat cong nghiep",
        "san van phong dien tich lon", "shophouse",
        # with diacritics
        "bi\u1ec7t th\u1ef1 \u0111\u01a1n l\u1eadp", "shophouse m\u1eb7t \u0111\u01b0\u1eddng l\u1edbn",
        "shophouse m\u1eb7t \u0111\u01b0\u1eddng", "qu\u1ef9 \u0111\u1ea5t c\u00f4ng nghi\u1ec7p",
        "s\u00e0n v\u0103n ph\u00f2ng di\u1ec7n t\u00edch l\u1edbn"
    ],
    "prime_locations": [
        "qu\u1eadn 1", "quan 1", "ven s\u00f4ng", "ven song",
        "vinhomes ocean park", "vinhomes oceanpark",
        "ph\u00fa m\u1ef9 h\u01b0ng", "phu my hung"
    ],
    "vip_profiles": [
        "ch\u1ee7 doanh nghi\u1ec7p", "chu doanh nghiep",
        "nh\u00e0 \u0111\u1ea7u t\u01b0 chuy\u00ean nghi\u1ec7p", "nha dau tu chuyen nghiep",
        "mua s\u1ec9", "mua si", "mua s\u1ed1 l\u01b0\u1ee3ng l\u1edbn", "mua so luong lon"
    ],
    "urgency_transparency": [
        "ph\u00e1p l\u00fd chu\u1ea9n 100%", "phap ly chuan 100%",
        "s\u1ed5 h\u1ed3ng ri\u00eang", "so hong rieng",
        "g\u1eb7p tr\u1ef1c ti\u1ebfp ch\u1ee7 \u0111\u1ea7u t\u01b0", "gap truc tiep chu dau tu"
    ]
}

JUNK_KEYWORDS_VI = {
    "no_need": [
        "nh\u1ea7m s\u1ed1", "nham so", "kh\u00f4ng c\u00f3 nhu c\u1ea7u", "khong co nhu cau",
        "d\u1eef li\u1ec7u c\u0169", "du lieu cu", "nh\u1ea7m ng\u00e0nh", "nham nganh"
    ],
    "uncooperative": [
        "h\u1ecfi gi\u00e1 cho vui", "hoi gia cho vui",
        "ch\u01b0a c\u00f3 \u00fd \u0111\u1ecbnh mua", "chua co y dinh mua",
        "th\u00e1i \u0111\u1ed9 kh\u00f4ng h\u1ee3p t\u00e1c", "thai do khong hop tac"
    ],
    "spam_adv": [
        "b\u1ea3o hi\u1ec3m", "bao hiem", "vay v\u1ed1n", "vay von",
        "m\u1eddi ch\u00e0o d\u1ecbch v\u1ee5", "moi chao dich vu",
        "qu\u1ea3ng c\u00e1o", "quang cao"
    ],
    "contact_issue": [
        "thu\u00ea bao", "thue bao", "kh\u00f4ng b\u1eaft m\u00e1y", "khong bat may",
        "kh\u00f4ng nh\u1ea5c m\u00e1y", "khong nhac may",
        "g\u1ecdi nhi\u1ec1u l\u1ea7n", "goi nhieu lan",
        "kh\u00f4ng ph\u1ea3n h\u1ed3i zalo", "khong phan hoi zalo"
    ]
}

# ---------------------------------------------------------------------------
# Text preprocessing
# ---------------------------------------------------------------------------
def remove_accents(input_str: str) -> str:
    if not input_str:
        return ""
    nfkd = unicodedata.normalize("NFKD", input_str)
    result = "".join(c for c in nfkd if not unicodedata.combining(c))
    return result.replace("\u0111", "d").replace("\u0110", "D")


def normalize_text(text: str):
    if not text:
        return "", ""
    text_lower = text.strip().lower()
    text_no_acc = remove_accents(text_lower)
    return text_lower, text_no_acc


# ---------------------------------------------------------------------------
# Scoring logic
# ---------------------------------------------------------------------------
def evaluate_lead(row: dict):
    """
    Score a lead based on description.
    Returns: (score, category, reasoning, vip_matched, junk_matched)
    """
    nhu_cau = row.get("nhu_cau_mo_ta", "") or ""
    if not nhu_cau.strip():
        return BASE_SCORE, "Bình thường", "Không có mô tả nhu cầu.", [], []

    text_lower, text_no_acc = normalize_text(nhu_cau)

    vip_matched = []
    junk_matched = []

    # --- VIP criteria (+50) ---

    # 1a. Large budget (>= 20 billion VND)
    budget_found = False
    budget_reason = ""
    for match in re.findall(r"(\d+(?:[.,]\d+)?)\s*(?:t\u1ef7|ty)", text_lower):
        try:
            val = float(match.replace(",", "."))
            if val >= 20.0:
                budget_found = True
                budget_reason = f"Ngan sach lon ({match} ty)"
                break
        except ValueError:
            continue

    for kw_list in [VIP_KEYWORDS_VI["large_budget_terms"], VIP_KEYWORDS["large_budget_terms"]]:
        for kw in kw_list:
            if kw in text_lower or kw in text_no_acc:
                budget_found = True
                budget_reason = f"Tu khoa ngan sach VIP ({kw})"
                break
        if budget_found:
            break

    if budget_found:
        vip_matched.append(budget_reason)

    # 1b. Luxury property type
    for kw_list in [VIP_KEYWORDS_VI["luxury_types"], VIP_KEYWORDS["luxury_types"]]:
        for kw in kw_list:
            if kw in text_lower or kw in text_no_acc:
                vip_matched.append(f"Loai hinh cao cap ({kw})")
                break
        else:
            continue
        break

    # 1c. Prime location
    has_q1 = bool(re.search(r"\b(qu\u1eadn\s*1|quan\s*1|q1|q\.1)\b", text_lower))
    if has_q1:
        vip_matched.append("Vi tri dac dia (Quan 1)")
    else:
        for kw_list in [VIP_KEYWORDS_VI["prime_locations"], VIP_KEYWORDS["prime_locations"]]:
            for kw in kw_list:
                if kw in ["qu\u1eadn 1", "quan 1"]:
                    continue
                if kw in text_lower or kw in text_no_acc:
                    vip_matched.append(f"Vi tri dac dia ({kw})")
                    break
            else:
                continue
            break

    # 1d. VIP customer profile
    for kw_list in [VIP_KEYWORDS_VI["vip_profiles"], VIP_KEYWORDS["vip_profiles"]]:
        for kw in kw_list:
            if kw in text_lower or kw in text_no_acc:
                vip_matched.append(f"Doi tuong khach hang VIP ({kw})")
                break
        else:
            continue
        break

    # 1e. Transparency / urgency
    for kw_list in [VIP_KEYWORDS_VI["urgency_transparency"], VIP_KEYWORDS["urgency_transparency"]]:
        for kw in kw_list:
            if kw in text_lower or kw in text_no_acc:
                vip_matched.append(f"Minh bach / Cap thiet ({kw})")
                break
        else:
            continue
        break

    # --- JUNK criteria (-50) ---

    # 2a. Unrealistic request (Quan 1 at very low price)
    unrealistic_found = False
    unrealistic_reason = ""
    if has_q1:
        low_price_kws = ["1 ty", "2 ty", "1-2 ty", "duoi 2 ty", "vai tram trieu", "may tram trieu",
                         "1 t\u1ef7", "2 t\u1ef7", "1-2 t\u1ef7", "d\u01b0\u1edbi 2 t\u1ef7",
                         "v\u00e0i tr\u0103m tri\u1ec7u", "m\u1ea5y tr\u0103m tri\u1ec7u"]
        for lp in low_price_kws:
            if lp in text_lower or lp in text_no_acc:
                unrealistic_found = True
                unrealistic_reason = f"Yeu cau phi thuc te (Quan 1 gia re: '{lp}')"
                break
        if not unrealistic_found:
            for m in re.findall(r"(\d+(?:[.,]\d+)?)\s*(?:t\u1ef7|ty)", text_lower):
                try:
                    val = float(m.replace(",", "."))
                    if val <= 2.5:
                        unrealistic_found = True
                        unrealistic_reason = f"Yeu cau phi thuc te (Quan 1 gia qua thap: {m} ty)"
                        break
                except ValueError:
                    continue

    if unrealistic_found:
        junk_matched.append(unrealistic_reason)

    # 2b-2e. Keyword junk checks
    for cat_key, kw_vi, kw_plain, label in [
        ("no_need",       JUNK_KEYWORDS_VI["no_need"],       JUNK_KEYWORDS["no_need"],       "Khong co nhu cau"),
        ("uncooperative", JUNK_KEYWORDS_VI["uncooperative"], JUNK_KEYWORDS["uncooperative"], "Khong thien chi"),
        ("spam_adv",      JUNK_KEYWORDS_VI["spam_adv"],      JUNK_KEYWORDS["spam_adv"],      "Spam / Quang cao"),
        ("contact_issue", JUNK_KEYWORDS_VI["contact_issue"], JUNK_KEYWORDS["contact_issue"], "Lien lac loi"),
    ]:
        matched = False
        for kw_list in [kw_vi, kw_plain]:
            for kw in kw_list:
                if kw in text_lower or kw in text_no_acc:
                    junk_matched.append(f"{label} ({kw})")
                    matched = True
                    break
            if matched:
                break

    # --- Final score ---
    vip_pts = VIP_BONUS if vip_matched else 0
    junk_pts = JUNK_PENALTY if junk_matched else 0
    final_score = max(0, min(100, BASE_SCORE + vip_pts + junk_pts))

    if final_score >= 100:
        category = "VIP"
        reason = f"Cộng 50 điểm (VIP): {', '.join(vip_matched)}"
        if junk_matched:
            reason += f" | Dấu hiệu tiêu cực: {', '.join(junk_matched)}"
    elif final_score <= 0:
        category = "Rác"
        reason = f"Trừ 50 điểm (Rác): {', '.join(junk_matched)}"
        if vip_matched:
            reason += f" | Có từ khóa VIP: {', '.join(vip_matched)}"
    else:
        category = "Bình thường"
        if vip_matched and junk_matched:
            reason = f"Giữ nguyên 50 điểm (Bình thường): Giao thoa VIP + Rác"
        else:
            reason = "Giữ nguyên 50 điểm (Bình thường): Khách hàng tầm trung."

    return final_score, category, reason, vip_matched, junk_matched


def get_gspread_client():
    """
    Xác thực và trả về đối tượng gspread client.
    Hỗ trợ:
    1. Service Account JSON file (ưu tiên 'gg-cloud-key-json.json' tiếp theo là 'service_account.json')
    2. OAuth token.json đã lưu tại ~/.config/ai_audit/token.json
    """
    import json
    import gspread
    from google.oauth2 import service_account
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request as GRequest

    SCOPES = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]

    # 1. Thử Service Account từ file 'gg-cloud-key-json.json' hoặc 'service_account.json'
    # Tìm theo absolute path dựa vào vị trí của file script này (không phụ thuộc CWD)
    script_dir = Path(__file__).parent.resolve()
    for key_file in ["gg-cloud-key-json.json", "service_account.json"]:
        # Tìm trong thư mục script trước, sau đó mới thử CWD
        for search_dir in [script_dir, Path.cwd()]:
            key_path = search_dir / key_file
            if key_path.exists():
                try:
                    with open(key_path, encoding="utf-8") as f:
                        key_data = json.load(f)
                    if key_data.get("type") == "service_account":
                        creds = service_account.Credentials.from_service_account_info(key_data, scopes=SCOPES)
                        return gspread.authorize(creds)
                except Exception:
                    pass

    # 2. Thử OAuth token từ các vị trí khả dụng
    token_locations = [
        Path.home() / ".config" / "ai_audit" / "token.json",
        Path("token.json"),
        Path(__file__).parent.resolve() / "token.json"
    ]
    
    TOKEN_PATH = None
    for loc in token_locations:
        if loc.exists():
            TOKEN_PATH = loc
            break

    if TOKEN_PATH:
        try:
            creds = Credentials.from_authorized_user_file(str(TOKEN_PATH), SCOPES)
            if not creds.valid:
                if creds.expired and creds.refresh_token:
                    creds.refresh(GRequest())
                    with open(TOKEN_PATH, "w") as f:
                        f.write(creds.to_json())
                else:
                    raise RuntimeError(
                        "OAuth token đã hết hạn và không thể làm mới. "
                        "Vui lòng chạy lại bước xác thực OAuth để lấy token mới."
                    )
            return gspread.authorize(creds)
        except RuntimeError:
            raise
        except Exception as e:
            raise RuntimeError(
                f"Lỗi khi tải OAuth token từ {TOKEN_PATH}:\n{e}\n"
                "Vui lòng xoá token.json và chạy lại xác thực OAuth."
            ) from e

    # Tạo danh sách các đường dẫn đã tìm để hiển thị thông báo lỗi chi tiết
    searched_paths = "\n".join([f"  • {str(p)}" for p in token_locations])
    raise RuntimeError(
        "Không tìm thấy thông tin xác thực Google.\n"
        "Vui lòng đặt một trong các tệp sau:\n"
        "  • File key của Service Account đặt tại thư mục dự án: gg-cloud-key-json.json hoặc service_account.json\n"
        f"  • File token OAuth cá nhân tại một trong các đường dẫn:\n{searched_paths}"
    )



# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------
def _get_worksheet_by_gid(sh, gid: str):
    """Trả về worksheet khớp gid. Nếu không tìm thấy, trả về sheet đầu tiên."""
    for ws in sh.worksheets():
        if str(ws.id) == str(gid):
            return ws
    return sh.get_worksheet(0)


def load_data(source: str) -> list:
    """
    Load data from Google Sheets URL or local CSV/Excel file.
    Returns list of dicts.
    """
    if source.startswith("http://") or source.startswith("https://"):
        if "docs.google.com/spreadsheets" in source:
            # Trích xuất gid (tab cụ thể) nếu có trong URL
            gid_match = re.search(r"[#&?]gid=(\d+)", source)
            gid = gid_match.group(1) if gid_match else None

            # 1. Thử tải có xác thực qua gspread (hỗ trợ Sheet riêng tư được share)
            auth_error = None
            try:
                gc = get_gspread_client()
                # Nếu xác thực thành công, mọi lỗi từ đây là lỗi quyền truy cập thực sự
                try:
                    sh = gc.open_by_url(source)
                    ws = _get_worksheet_by_gid(sh, gid) if gid else sh.get_worksheet(0)
                    all_values = ws.get_all_values()
                    if all_values:
                        headers = all_values[0]
                        rows = []
                        for row in all_values[1:]:
                            row_dict = {}
                            for idx, h in enumerate(headers):
                                if h:
                                    row_dict[h] = row[idx] if idx < len(row) else ""
                            rows.append(row_dict)
                        return rows
                except Exception as sheet_err:
                    # Đã xác thực nhưng không mở được sheet → báo lỗi thay vì fallback ngầm
                    raise RuntimeError(
                        f"Đã xác thực Google thành công nhưng không thể mở Sheet.\n"
                        f"→ Kiểm tra rằng Sheet đã được chia sẻ cho tài khoản đang dùng.\n"
                        f"Chi tiết: {sheet_err}"
                    )
            except RuntimeError:
                # Lỗi quyền/sheet → ném lên để Streamlit hiển thị
                raise
            except Exception as auth_err:
                # Không có credentials → thử fallback CSV công khai
                auth_error = auth_err

            # 2. Fallback tải công khai qua requests export CSV
            match = re.search(r"/d/([a-zA-Z0-9-_]+)", source)
            if match:
                sheet_id = match.group(1)
                gid_str = f"&gid={gid}" if gid else ""
                csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv{gid_str}"
            else:
                csv_url = source

            try:
                resp = requests.get(csv_url, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
                resp.raise_for_status()
                csv_data = resp.content.decode("utf-8-sig")
                reader = csv.DictReader(io.StringIO(csv_data))
                return list(reader)
            except requests.HTTPError as csv_err:
                if csv_err.response is not None and csv_err.response.status_code in (401, 403):
                    raise RuntimeError(
                        "Sheet này là riêng tư và không thể tải công khai.\n"
                        "→ Hãy chia sẻ Sheet với tài khoản Google đã xác thực, "
                        "hoặc đặt quyền 'Anyone with the link can view'."
                    ) from csv_err
                raise

            resp = requests.get(csv_url, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
            resp.raise_for_status()
            return []

        resp = requests.get(source, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
        resp.raise_for_status()
        csv_data = resp.content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(csv_data))
        return list(reader)
    else:
        path = Path(source)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {source}")
        if path.suffix.lower() == ".csv":
            with open(path, encoding="utf-8-sig") as f:
                return list(csv.DictReader(f))
        elif path.suffix.lower() in [".xlsx", ".xls"]:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows = []
            for r in range(2, ws.max_row + 1):
                row_dict = {}
                for c_idx, h in enumerate(headers, 1):
                    if h:
                        row_dict[h] = ws.cell(row=r, column=c_idx).value
                rows.append(row_dict)
            return rows
        else:
            raise ValueError("Unsupported format. Use .csv, .xlsx or Google Sheet URL.")


# ---------------------------------------------------------------------------
# Excel report generation
# ---------------------------------------------------------------------------
def create_excel_report(scored_leads: list, integrity_log: list, output_path) -> None:
    """
    Create a professional Excel report with 3 sheets:
      1. Summary_Dashboard
      2. Lead_Scores_Details
      3. Data_Integrity_Log
    output_path can be str, Path, or BytesIO buffer.
    """
    FONT = "Segoe UI"

    fill_navy = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_kpi  = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    fill_vip  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_junk = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    fill_norm = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    fill_err  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    thin  = Side(border_style="thin",   color="D9D9D9")
    thick = Side(border_style="medium", color="1F4E79")
    dbl   = Side(border_style="double", color="1F4E79")

    b_cell  = Border(left=thin, right=thin, top=thin, bottom=thin)
    b_kpi   = Border(left=thin, right=thin, top=thin, bottom=thick)
    b_total = Border(top=thin, bottom=dbl)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ---- Sheet 1: Summary Dashboard ----------------------------------------
    ws1 = wb.create_sheet("Summary_Dashboard")

    ws1["A1"] = "BÁO CÁO PHÂN TÍCH CHẤT LƯỢNG KHÁCH HÀNG TIỀM NĂNG"
    ws1["A1"].font = Font(name=FONT, size=16, bold=True, color="1F4E79")
    ws1["A2"] = f"Hệ thống chấm điểm tự động | Thời gian xuất: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws1["A2"].font = Font(name=FONT, size=10, italic=True, color="595959")
    ws1.row_dimensions[1].height = 26
    ws1.row_dimensions[2].height = 18

    total   = len(scored_leads)
    n_vip   = sum(1 for r in scored_leads if r["__score"][1] == "VIP")
    n_norm  = sum(1 for r in scored_leads if r["__score"][1] == "Bình thường")
    n_junk  = sum(1 for r in scored_leads if r["__score"][1] == "Rác")
    avg_sc  = round(sum(r["__score"][0] for r in scored_leads) / total, 1) if total else 0

    kpi_data = [
        ("B", "TỔNG SỐ KHÁCH",    total,  "#,##0"),
        ("C", "KHÁCH VIP",        n_vip,  "#,##0"),
        ("D", "BÌNH THƯỜNG",      n_norm, "#,##0"),
        ("E", "KHÁCH RÁC",        n_junk, "#,##0"),
        ("F", "ĐIỂM TRUNG BÌNH",  avg_sc, "0.0"),
    ]
    for col, lbl, val, fmt in kpi_data:
        lc = ws1[f"{col}4"]
        lc.value = lbl
        lc.font  = Font(name=FONT, size=9, bold=True, color="595959")
        lc.fill  = fill_kpi
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border = Border(left=thin, right=thin, top=thin)

        vc = ws1[f"{col}5"]
        vc.value  = val
        vc.font   = Font(name=FONT, size=22, bold=True, color="1F4E79")
        vc.fill   = fill_kpi
        vc.number_format = fmt
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = b_kpi

    ws1.row_dimensions[4].height = 18
    ws1.row_dimensions[5].height = 34

    ws1["B7"] = "BẢNG PHÂN PHỐI CƠ CẤU KHÁCH HÀNG"
    ws1["B7"].font = Font(name=FONT, size=12, bold=True, color="1F4E79")
    ws1.merge_cells("B7:C7")
    ws1.row_dimensions[7].height = 22

    for col_lbl, heading in [("B", "Nhóm Khách Hàng"), ("C", "Số Lượng")]:
        c = ws1[f"{col_lbl}8"]
        c.value = heading
        c.font  = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill  = fill_navy
        c.alignment = Alignment(horizontal="center" if col_lbl == "C" else "left", vertical="center")
    ws1.row_dimensions[8].height = 22

    for row_i, (cat, val) in enumerate([("VIP (100 điểm)", n_vip), ("Bình thường (50 điểm)", n_norm), ("Rác (0 điểm)", n_junk)], 9):
        b = ws1[f"B{row_i}"]
        c = ws1[f"C{row_i}"]
        b.value = cat;   b.font = Font(name=FONT, size=10); b.border = b_cell
        c.value = val;   c.font = Font(name=FONT, size=10); c.border = b_cell
        c.number_format = "#,##0"
        c.alignment = Alignment(horizontal="right")
        if row_i % 2 == 0:
            b.fill = fill_zebra; c.fill = fill_zebra

    tot_row = 12
    for col_letter, val in [("B", "Tổng cộng"), ("C", "=SUM(C9:C11)")]:
        cell = ws1[f"{col_letter}{tot_row}"]
        cell.value  = val
        cell.font   = Font(name=FONT, size=10, bold=True)
        cell.border = b_total
        if col_letter == "C":
            cell.alignment = Alignment(horizontal="right")

    for col_letter, w in [("A", 3), ("B", 28), ("C", 16), ("D", 18), ("E", 18), ("F", 18)]:
        ws1.column_dimensions[col_letter].width = w

    # Pie chart
    try:
        from openpyxl.chart import PieChart, Reference
        pie = PieChart()
        labels = Reference(ws1, min_col=2, min_row=9, max_row=11)
        data   = Reference(ws1, min_col=3, min_row=8, max_row=11)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title  = "Cơ Cấu Khách Hàng Tiềm Năng"
        pie.width  = 16
        pie.height = 10
        ws1.add_chart(pie, "B14")
    except Exception:
        pass

    # ---- Sheet 2: Lead_Scores_Details --------------------------------------
    ws2 = wb.create_sheet("Lead_Scores_Details")
    headers2 = ["Mã Lead", "Tên Khách Hàng", "Số Điện Thoại", "Mô Tả Nhu Cầu",
                "Điểm Cơ Bản", "Điểm Cộng (VIP)", "Điểm Trừ (Rác)", "Điểm Cuối",
                "Phân Loại", "Giải Trình Lý Do"]

    for ci, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font  = Font(name=FONT, size=11, bold=True, color="FFFFFF")
        cell.fill  = fill_navy
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = b_cell
    ws2.row_dimensions[1].height = 30

    for ri, lead in enumerate(scored_leads, 2):
        score, cat, reason, vip_m, junk_m = lead["__score"]

        ws2.cell(ri, 1, lead.get("id", f"L{ri-1}")).alignment = Alignment(horizontal="center")
        ws2.cell(ri, 2, lead.get("ten_khach", "Ẩn danh"))
        ws2.cell(ri, 3, lead.get("sdt", "N/A")).alignment = Alignment(horizontal="center")
        ws2.cell(ri, 4, lead.get("nhu_cau_mo_ta", "")).alignment = Alignment(wrap_text=True)
        ws2.cell(ri, 5, BASE_SCORE).number_format = "#,##0"
        ws2.cell(ri, 6, VIP_BONUS if vip_m else 0).number_format = "#,##0"
        ws2.cell(ri, 7, JUNK_PENALTY if junk_m else 0).number_format = "#,##0"

        sc = ws2.cell(ri, 8, score)
        sc.font = Font(name=FONT, size=10, bold=True)
        sc.number_format = "#,##0"
        sc.alignment = Alignment(horizontal="center")

        cc = ws2.cell(ri, 9, cat)
        cc.alignment = Alignment(horizontal="center")
        if cat == "VIP":
            cc.fill = fill_vip;  cc.font = Font(name=FONT, size=10, bold=True, color="375623")
        elif cat == "Rác":
            cc.fill = fill_junk; cc.font = Font(name=FONT, size=10, bold=True, color="C65911")
        else:
            cc.fill = fill_norm; cc.font = Font(name=FONT, size=10, bold=True, color="1F4E79")

        ws2.cell(ri, 10, reason).alignment = Alignment(wrap_text=True)

        for ci in range(1, len(headers2) + 1):
            c = ws2.cell(ri, ci)
            c.border = b_cell
            if ci not in [8, 9]:
                c.font = Font(name=FONT, size=10)
            if ri % 2 == 0 and ci != 9:
                c.fill = fill_zebra

        ws2.row_dimensions[ri].height = 38

    col_widths2 = {"A": 12, "B": 22, "C": 16, "D": 45, "E": 14,
                   "F": 15, "G": 14, "H": 12, "I": 14, "J": 45}
    for col_l, w in col_widths2.items():
        ws2.column_dimensions[col_l].width = w

    # ---- Sheet 3: Data_Integrity_Log ---------------------------------------
    ws3 = wb.create_sheet("Data_Integrity_Log")
    headers3 = ["Dòng Sheet", "Mã ID", "Tên Khách Hàng", "Số Điện Thoại",
                "Trường Bị Lỗi", "Chi Tiết Lỗi", "Hướng Xử Lý"]

    for ci, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=ci, value=h)
        cell.font  = Font(name=FONT, size=11, bold=True, color="FFFFFF")
        cell.fill  = fill_navy
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = b_cell
    ws3.row_dimensions[1].height = 26

    if integrity_log:
        for ri, err in enumerate(integrity_log, 2):
            ws3.cell(ri, 1, err.get("row_num")).alignment = Alignment(horizontal="center")
            ws3.cell(ri, 2, err.get("id")).alignment = Alignment(horizontal="center")
            ws3.cell(ri, 3, err.get("ten_khach", ""))
            ws3.cell(ri, 4, err.get("sdt", "")).alignment = Alignment(horizontal="center")
            ws3.cell(ri, 5, err.get("field")).alignment = Alignment(horizontal="center")

            ec = ws3.cell(ri, 6, err.get("error_msg", ""))
            ec.font = Font(name=FONT, size=10, color="9C0006")
            ec.fill = fill_err

            ws3.cell(ri, 7, err.get("action", ""))

            for ci in range(1, len(headers3) + 1):
                c = ws3.cell(ri, ci)
                c.border = b_cell
                if ci != 6:
                    c.font = Font(name=FONT, size=10)
                if ri % 2 == 0 and ci != 6:
                    c.fill = fill_zebra
            ws3.row_dimensions[ri].height = 24
    else:
        ws3.merge_cells("A2:G2")
        ok = ws3.cell(2, 1, "Tuyệt vời! Không phát hiện lỗi chất lượng dữ liệu nào.")
        ok.font = Font(name=FONT, size=11, bold=True, color="375623")
        ok.fill = fill_vip
        ok.alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[2].height = 34
        for ci in range(1, len(headers3) + 1):
            ws3.cell(2, ci).border = b_cell

    for col_l, w in [("A", 12), ("B", 14), ("C", 24), ("D", 18), ("E", 18), ("F", 40), ("G", 40)]:
        ws3.column_dimensions[col_l].width = w

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Optional: write scores back to Google Sheet (requires gspread + credentials)
# ---------------------------------------------------------------------------
def write_scores_to_gspread(scored_leads: list, source_url: str) -> None:
    try:
        import gspread
    except ImportError:
        raise RuntimeError("gspread is not installed. Run: pip install gspread google-auth")

    gc = get_gspread_client()
    sh = gc.open_by_url(source_url)
    ws = sh.get_worksheet(0)
    headers = ws.row_values(1)

    new_cols = ["Final_Score", "Category", "Reasoning_Details"]
    col_indices = {}
    for col_name in new_cols:
        if col_name in headers:
            col_indices[col_name] = headers.index(col_name) + 1
        else:
            headers.append(col_name)
            col_indices[col_name] = len(headers)
            ws.update_cell(1, col_indices[col_name], col_name)

    updates = []
    for idx, lead in enumerate(scored_leads, 2):
        score, cat, reason, _, _ = lead["__score"]
        for col_name, val in [("Final_Score", score), ("Category", cat), ("Reasoning_Details", reason)]:
            updates.append({
                "range": gspread.utils.rowcol_to_a1(idx, col_indices[col_name]),
                "values": [[val]]
            })
    ws.batch_update(updates)
